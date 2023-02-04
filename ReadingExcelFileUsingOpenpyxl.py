import pandas as pd
import os
import xlrd
import xlrd3
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import pyautogui
import pywinauto


wrkbk = load_workbook("C:\\Automation\\TestCaseInputFile.xlsx")
# to identify the active sheet
sh = wrkbk.active
# get the value of row 2 and column7
c1=sh.cell(row=2,column=7)
print(c1.value)
